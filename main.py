import os
import pathlib
import uuid

from PIL import Image
import numpy as np
import openpyxl
from openpyxl.utils.cell import get_column_letter
import streamlit as st


@st.cache_data(persist="disk")
def transform_image_to_xlsx(image_path: str, default_cols: int = 50, is_rescale=False) -> (str, str):
    """
    Transforms an image into an xlsx file where each cell corresponds to a pixel of the image.
    Each cell is filled with a color corresponding to the RGB value of the corresponding pixel.

    Args:
        image_path (str): The path to the image file.
        default_cols (int, optional): The number of columns to use if `is_scale` is True. Defaults to 50.
        is_rescale (bool, optional): Whether to scale the image to fit `default_cols` columns. Defaults to False.

    Returns:
        file_path (str): xlsx file full path.
        file_name (str): xlsx file unique file name.
    """
    # Always check for output directory
    pathlib.Path("output").mkdir(parents=True, exist_ok=True)

    # Open image
    with Image.open(image_path) as image:
        if is_rescale:
            # Scale image if requested
            width, height = image.size
            image_scale = default_cols / width
            scaled_size = (int(width * image_scale), int(height * image_scale))
            image = image.resize(scaled_size, resample=Image.Resampling.LANCZOS)

        # Convert image to RGB array
        rgb_array = np.array(image.convert("RGB"))
        width, height, _ = rgb_array.shape

        # Create xlsx workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Create list of fill colors for each pixel
        fill_colors = [f"{r:02X}{g:02X}{b:02X}" for r, g, b in rgb_array.reshape(-1, 3)]

        # Fill worksheet cells with appropriate color
        for x in range(0, width):
            column_letter = get_column_letter(x + 1)
            worksheet.column_dimensions[column_letter].width = 3
        
            for y in range(0, height):
                fill_color = fill_colors[x * height + y]
                cell = worksheet.cell(row=x + 1, column=y + 1)
                cell.fill = openpyxl.styles.PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        # Save workbook with unique filename
        file_name = f"{uuid.uuid4().hex[:16]}.xlsx"
        file_path = os.path.join("output", file_name)
        workbook.save(file_path)

        return (file_path, file_name)


if __name__ == "__main__":
    st.title("ikusel (📷💻📊)")
    st.markdown("[![GitHub Star](https://img.shields.io/github/stars/Kawaeee/ikusel)](https://github.com/Kawaeee/ikusel)")
    st.write("Simple web-based tool that lets users upload an image and convert it to an Excel spreadsheet (.xlsx) file!")

    uploaded_file = st.file_uploader(label="Upload Image..", type=["png", "jpg", "jpeg"], label_visibility="hidden")
    left, right = st.columns(2)

    with left:
        is_rescale = st.checkbox("Rescale", value=True)

        if is_rescale:
            default_cols = st.number_input("How many columns that would fit your image?", value=50, min_value=25, max_value=200)
        else:
            default_cols = 0
            is_rescale = False

    if uploaded_file is not None:
        with left:
            submit_button = st.button(label="Submit")
            if submit_button:
                file_path, file_name = transform_image_to_xlsx(uploaded_file, default_cols, is_rescale)
                with open(file_path, "rb") as f:
                    download_button = st.download_button(label="Download", data=f, file_name=file_name)
        with right:
            st.image(uploaded_file, use_column_width="auto", caption="User uploaded image")
